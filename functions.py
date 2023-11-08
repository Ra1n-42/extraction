import os
import re
import time
import datetime
from datetime import datetime as dt
import json
import threading
from openpyxl import load_workbook

class AttrDict(dict):
    def __getattr__(self, key):
        if key in self:
            return self[key]
        raise AttributeError(f"'AttrDict' object has no attribute '{key}'")


config_file_path = 'config.json'
try:
    with open(config_file_path, encoding="utf-8") as f:
        config = AttrDict(json.load(f))

except FileNotFoundError:
    print(f"Datei nicht gefunden:")
    print()
    print(f"Die {config_file_path} Datei muss sich im selben Verzeichnis wie die .exe befinden.")
    print()
    print("Drück eine Taste um zum beenden")


EXTRACTED_PATH = config.source_path
EXTRACTED_FILE = config.file_path


#  Funktionen

def getListOfOverviewIDs(sheet):
    header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    result = []
    for i, s in enumerate(header_row):
        # extrahiert die Kostenstelle aus EXTRACTED_FILE
        accounting_number = extract_long_number_from_Ac(s)
        if accounting_number:
            result.append((i, accounting_number.ljust(config.max_laenge_kostenstelle, "0")))
        else:
            pass  # print(f'{s} die kostenstelle(n) Nummer konnten nicht extrahiert werden, das sind die "kostenstellen" ohne Nummern
    return result


def find_longest_string(tuple_list):
    max_len = 0
    max_str = ""

    for _, _1 in tuple_list:
        val = str(_1)
        if len(val) > max_len:
            max_len = len(val)
            max_str = val

    return max_str, max_len


def find_cell(sheet, date, kostenstelle):
    '''
    :function > es soll die Zelle oder auch koordinate ausgeben
    :param sheet: aktelle sheet, mit sheet.title etc., glaube ist ein Obj.
    :param date: datum aus der FAKE_DB= ist Auszug aus den files, ist ein string
    :param kostenstelle: aus der data/FAKE_DB, string oder int
    :return: koordinate der Zelle
    '''
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=1):
        for cell in row:
            splitteddates = str(cell.value).split(' ')
            if formatUsDatesToEU(splitteddates[0]) == date and len(splitteddates) > 1:

                spaltenzahl_und_kostenstellen = getListOfOverviewIDs(sheet)
                config.max_laenge_kostenstelle = find_longest_string(spaltenzahl_und_kostenstellen)[1]
                index = find_short_num(
                    str(kostenstelle).ljust(config.max_laenge_kostenstelle, "0"),
                    spaltenzahl_und_kostenstellen)

                if index:
                    column_number = index + 1
                    row_number = cell.row
                    return sheet.cell(row=row_number, column=column_number).coordinate
                else:
                    config.meldungen.append(
                        f'Die "Kostenstelle" : {kostenstelle} aus Tabelle: "{sheet.title}", wurde in der "{EXTRACTED_FILE}" nicht gefunden.')
    return None


def update_to_file(sheet_names, sheet, sheet_name):
    __update_count = 0
    for _ in config.data:
        for date, entry in _.items():
            if entry['Name'] in sheet_names:
                if sheet_name == entry['Name']:
                    coord = find_cell(sheet, date, entry['BaustellenNr'])  # koordinaten

                    if coord:
                        sheet[coord].value = entry["AB"]
                        __update_count += 1
    return __update_count


################################################### END


# def getConstrNumber(sheet):
#     __tmp = ''
#     if sheet["D25"].value:
#         for coord in AttrDict(config.positionen).Koststelle:
#             value = sheet[coord].value or 0
#             __tmp += str(value)
#         return int(__tmp[:config.max_laenge_kostenstelle])
#     else:
#         return 0


def getConstrNumber(sheet):
    if sheet["D25"].value:
        values = []
        for coord in AttrDict(config.positionen).Koststelle:
            value = sheet[coord].value or 0
            values.append(str(value))

        return int(''.join(values)[:config.max_laenge_kostenstelle])
    return None



def getStartTime(sheet) -> dict:
    __tmp = str(sheet[AttrDict(config.positionen).MontagDatum].value)
    date_object = dt.strptime(__tmp, "%Y-%m-%d %H:%M:%S")
    year = date_object.year
    month = date_object.month
    day = date_object.day
    return AttrDict({'year': int(year), 'month': int(month), 'day': int(day)})


def getDataFromWeekWorkTime(sheet, file) -> bool:
    work_time = AttrDict(config.positionen).Arbeitszeit
    for cells in work_time:
        kostenstelle = getConstrNumber(sheet)
        if kostenstelle:
            hours = sheet[cells[0]].value or 0
            minutes = sheet[cells[1]].value or 0
            zeit = getStartTime(sheet)
            dateofentrystring = datetime.date(zeit.year, zeit.month, zeit.day) + datetime.timedelta(
                days=work_time.index(cells))
            datum = dateofentrystring.strftime('%d.%m.%Y')
            name = getName(sheet)

            tmp_liste = re.split(',|\.', str(hours))
            if len(tmp_liste) > 1:
                if tmp_liste[1].isdigit():
                    hours, minutes = tmp_liste[0], tmp_liste[1]
                else:
                    hours, minutes = tmp_liste[0], minutes

            arbeitszeit = str(hours) + '.' + str(minutes)
            config.data.append(
                {datum: {'Name': name, 'AB': float(arbeitszeit), 'BaustellenNr': kostenstelle, 'filename': file}})
            return True
        else:
            config.notexisting.append(
                f'Die "Baustelle"n Nummer, wurde links in der Tabelle "{sheet.title}" in der Datei "{file}", nicht gefunden.')
            return False

def extract_long_number_from_Ac(string):
    if string is not None:
        isint = str(string).split()[0]
        if isint[0].isdigit():
            return isint
    return None


def formatUsDatesToEU(string) -> str:
    lst = string.split('-')
    lst[0], lst[-1] = lst[-1], lst[0]
    return '.'.join(lst)


def find_short_num(long_num, num_list):
    for num_tuple in num_list:
        if num_tuple[1] == long_num:
            return num_tuple[0]
    return None


def getName(sheet):
    return sheet[AttrDict(config.positionen).Nachname].value



